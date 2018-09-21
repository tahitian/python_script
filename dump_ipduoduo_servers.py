import json
import sys
from openpyxl import Workbook
from openpyxl import load_workbook

def load_server_info():
    info = {}
    file_path = "%s/ipduoduo_servers.txt" % sys.path[0]
    file_path = "./ipduoduo_servers.txt"
    print('path: %s' % sys.path[0])
    # file_path = "%s/data/pptp/configs/bzy_servers.txt" % sys.path[0]
    try:
        fd = open(file_path)
        if fd:
            lines = fd.readlines()
        fd.close()
    except Exception as e:
        print("load_server_info: %r" % e)
    for line in lines:
        try:
            data = json.loads(line)
            # if not data["enabled"]:
            #     continue
            name = data["server"]
            rate = data["rate"]
            enabled = data['enabled']
            # mppe = data["mppe"]
            info[name] = {"rate": rate, "enabled": enabled}
        except Exception as e:
            print("load_server_info: %r" % e)
    return info

def dump_to_xslx(info):
    wb = Workbook()
    sheet = wb.active
    # wb = load_workbook('%s/ipduoduo_servers.xlsx' % sys.path[0])
    # sheet = wb['Sheet1']
    sheet['A1'] = 'server'
    sheet['B1'] = 'enabled'
    sheet['C1'] = 'rate'
    row = 2
    for name in info:
        sheet['A%d'%row] = name
        sheet['B%d'%row] = info[name]['enabled']
        sheet['C%d'%row] = info[name]['rate']
        row += 1
    wb.save('ipduoduo.xlsx')

if __name__ == '__main__':
    info = load_server_info()
    dump_to_xslx(info)