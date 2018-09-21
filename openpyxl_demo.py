import os
from openpyxl import Workbook
from openpyxl import load_workbook

path = os.path.dirname(os.path.abspath(__file__))

# wb = Workbook()
# sheet = wb.active
# sheet['A1'] = 'hello'
# wb.save('%s/test.xlsx' % path)

wb = load_workbook('test.xlsx')
sheet = wb['Sheet']
n = sheet['A1'].value
print(n)
sheet['A1'] = 'world'
wb.save('%s/test.xlsx' % path)